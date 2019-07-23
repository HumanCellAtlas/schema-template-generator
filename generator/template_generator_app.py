#!/usr/bin/env python
import sys

import os
import tempfile

import yaml
from flask import Flask, Markup, flash, request, render_template, redirect, url_for, make_response, send_file
from flask_cors import CORS, cross_origin
import logging
import configparser
import datetime
from openpyxl import load_workbook
from ingest.template.schema_template import SchemaTemplate, UnknownKeyException
from ingest.template.spreadsheet_builder import SpreadsheetBuilder

EXCLUDED_PROPERTIES = ["describedBy", "schema_version", "schema_type", "provenance"]

INGEST_API_URL = "http://api.ingest.{env}.data.humancellatlas.org"

STATUS_LABEL = {
    'Valid': 'label-success',
    'Validating': 'label-info',
    'Invalid': 'label-danger',
    'Submitted': 'label-default',
    'Complete': 'label-default'
}

ALLOWED_EXTENSIONS = set(['yaml', 'yml', 'xls', 'xlsx'])
UPLOAD_FOLDER = 'tmp/yaml_file'

DEFAULT_STATUS_LABEL = 'label-warning'


HTML_HELPER = {
    'status_label': STATUS_LABEL,
    'default_status_label': DEFAULT_STATUS_LABEL
}

# Flask boiler plate
app = Flask(__name__, static_folder='static')
app.secret_key = 'cells'
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

logger = logging.getLogger(__name__)

# dictionary of schema to tab names as tab names can be hard to get hold of from the schema template library, esp for sub-tabs
DISPLAY_NAME_MAP = {}

CONFIG_FILE = ''

SCHEMA_TEMPLATE = {}

#function that takes an uploaded YAML file and renders it in the context of all the latest schemas
@app.route('/upload', methods=['POST'])
def upload_file():

    if 'yamlfile' not in request.files:
        flash('Warning! No file provided!')
        return redirect(url_for('index'))
    file = request.files['yamlfile']

    if file.filename == '':
        flash('Warning! File name blank!')
        return redirect(url_for('index'))
    if file and _allowed_file(file.filename):

        content = yaml.load(file.stream.read(), Loader=yaml.FullLoader)

        # load all properties from latest schemas
        all_properties = _process_schemas()

        # get all the properties from the file
        selected_schemas, selected_properties = _process_uploaded_file(content['tabs'])

        # tag all the properties from the file as preselected in the full properties list
        schema_properties = _preselect_properties(all_properties, selected_schemas, None, selected_properties)

        # return the list of pre-selected properties to be rendered
        return render_template('schemas.html', helper=HTML_HELPER, schemas=schema_properties)

# function that takes a YAML file and converts it straight to a spreadsheet without going via the property selection page
@app.route('/upload_yaml_to_xls', methods=['POST'])
def upload_generate():

    if 'yamlfile' not in request.files:
        flash('Warning! No file provided!')
        return redirect(url_for('index'))
    file = request.files['yamlfile']

    if file.filename == '':
        flash('Warning! File name blank!')
        return redirect(url_for('index'))
    if file and _allowed_file(file.filename):

        yaml_json = yaml.load(file.stream.read(), Loader=yaml.FullLoader)
        response = _generate_spreadsheet(yaml_json)
        return response


# function that loads all properties for all schemas - called either from home page or schemas preselection page
@app.route('/load_all', methods=['GET', 'POST'])
def load_full_schemas():
    all_properties = _process_schemas()

    response = request.form

    # if this function is called from the preselection page, there should be a list of pre-selected schemas
    selected_schemas = []
    if 'schema' in response:
        selected_schemas = response.getlist('schema')

    # if this function is called from the preselection page, there should be a list of pre-selected modules (references
    selected_references = []
    if 'reference' in response:
        selected_references = response.getlist('reference')
        for ref in selected_references:
            if ref.split(":")[1] in DISPLAY_NAME_MAP:
                if ref.split(":")[0] in selected_schemas:
                    selected_schemas.append(ref.split(":")[1])

    # process all schemas to tag any preselected ones for ticking
    schema_properties = _preselect_properties(all_properties, selected_schemas, selected_references, None)

    return render_template('schemas.html', helper=HTML_HELPER, schemas=schema_properties)

# function that loads schemas and modules (references) only for preselection
@app.route('/load_select', methods=['GET'])
def selectSchemas():

    # load the tab config and go through all the schemas
    tab_config = SCHEMA_TEMPLATE.get_tabs_config()

    unordered = {}
    for schema in tab_config.lookup('tabs'):
        schema_name = list(schema.keys())[0]

        properties = schema[schema_name]['columns']
        schema_title = schema[schema_name]['display_name']

        schema_structure = tab_config.lookup('meta_data_properties')[schema_name]

        references = _extract_references(properties, schema_name, schema_title, schema_structure)
        unordered[references["name"]] = references

    #make sure schemas and subschemas are presented in the order defined in the config file
    orderedReferences = []
    if 'ordering' in CONFIG_FILE:
        for key in CONFIG_FILE['ordering'].keys():
            if key in unordered.keys():
                orderedReferences.append(unordered[key])
            else:
                print(key + " is currently not a recorded property")

    return render_template('schema_selector.html', helper=HTML_HELPER, schemas=orderedReferences)

# get the index page
@app.route('/')
def index():
    return render_template('index.html', helper=HTML_HELPER)

# generate a spreadsheet or YAML file from the pre-selected schemas and properties
@app.route('/generate', methods=['POST'])
def generate_yaml():

    response = request.form

    # get the list of selected schemas and properties from the request
    selected_schemas = []
    if 'schema' in response:
        selected_schemas = response.getlist('schema')

    selected_properties = []
    if 'property' in response:
        selected_properties = response.getlist('property')

    # build a yaml-style structure of the selected schemas and properties
    pre_yaml = []
    for schema in selected_schemas:
        entry = {}
        tab = {}

        if DISPLAY_NAME_MAP[schema]:
            tab["display_name"] = DISPLAY_NAME_MAP[schema]
        else:
            tab["display_name"] = schema

        columns = []
        for prop in selected_properties:
            t = prop.split(':')[0]
            val = prop.split(':')[1]
            if schema == t:
                columns.append(val)
        tab["columns"] = columns
        entry[schema] = tab
        pre_yaml.append(entry)

    yaml_json = {}
    yaml_json["tabs"] = pre_yaml

    # to generate a yaml file, dump the data structure out to yaml format
    if request.form['submitButton'] == 'yaml':
        yaml_data = yaml.dump(yaml_json, default_flow_style=False)
        now = datetime.datetime.now()
        filename = "hca_yaml-" + now.strftime("%Y-%m-%dT%H-%M-%S") + ".yaml"
        response = make_response(yaml_data)
        response.headers.set('Content-Type', 'application/x-yaml')
        response.headers.set('Content-Disposition', 'attachment',
                             filename=filename)
        return response

    # to generate a spreadsheet, conver the yaml json format to spreadsheet
    elif request.form['submitButton'] == 'spreadsheet':

            response = _generate_spreadsheet(yaml_json)
            return response

# function to migrate an older spreadsheet (incl data) to the latest schema version
@app.route('/upload_xls', methods=['POST'])
def upload_spreadsheet():

    response = request.form

    if 'xlsfile' not in request.files:
        flash('Warning! No file provided!')
        return redirect(url_for('index'))
    file = request.files['xlsfile']

    if file.filename == '':
        flash('Warning! File name blank!')
        return redirect(url_for('index'))
    if file and _allowed_file(file.filename):
        # TO DO this is a hack to get past the 'No such file or directory' error but relies on directory being present - FIX!
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
        wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))

        # we no longer needs the schemas tab but I left this in for now
        if 'Schemas' in wb.sheetnames:
            schemas = wb['Schemas']
        elif 'schemas' in wb.sheetnames:
            schemas = wb['schemas']
        else:
           schemas = None

        latest_schemas = SCHEMA_TEMPLATE.get_schema_urls()

        tab_config = SCHEMA_TEMPLATE.get_tabs_config()
        tabs = wb.sheetnames

        # go through each schema in the tab config and if it's present in the old spreadsheet, migrate it
        # this approach is slightly inefficient as it ends up migrating schemas that are already at their latest version but it means
        # no schemas tab is required - in order to migrate only tabs that have actually changed, we'd need to know the schema version of
        # each tab in the old spreadsheet
        for schema in tab_config.lookup('tabs'):
            tab_name = schema[list(schema.keys())[0]]['display_name']

            if tab_name in tabs:
                schema_name = list(schema.keys())[0]

                for schema in latest_schemas:
                    if schema_name in schema:
                        print("Migrating " + tab_name)
                        _migrate_schema(wb, schema)

        # usual slightly complicated set-up to return the output spreadsheet
        with tempfile.NamedTemporaryFile('w+b', delete=False) as ssheet_file:
            temp_filename = ssheet_file.name

            now = datetime.datetime.now()
            export_filename = "hca_spreadsheet-" + now.strftime("%Y-%m-%dT%H-%M-%S") + "_migrated.xlsx"
            wb.save(temp_filename)
            wb.close()

            response = make_response(ssheet_file.read())
            response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response.headers.set('Content-Disposition', 'attachment',
                                 filename=export_filename)
            os.remove(temp_filename)
            return response

# convenience method to generate a spreadsheet from a pre-yaml data structure using the schema template library's spreadsheet generator
def _generate_spreadsheet(yaml_json):

    # I know this looks like an unused property but it somehow doesn't work if you don't declare it here first
    # slightly complicated set-up with named temp files to pass structures around and have a persistent file object that we can actually return to the browser
    temp_yaml_filename = ""
    with tempfile.NamedTemporaryFile('w', delete=False) as yaml_file:
        yaml.dump(yaml_json, yaml_file)
        temp_yaml_filename = yaml_file.name

    with tempfile.NamedTemporaryFile('w+b', delete=False) as ssheet_file:
        temp_filename = ssheet_file.name
        spreadsheet_builder = SpreadsheetBuilder(temp_filename, True)
        # TO DO currently automatically building WITH schemas tab - this should be customisable
        spreadsheet_builder.generate_workbook(tabs_template=temp_yaml_filename,
                                              schema_urls=SCHEMA_TEMPLATE.get_schema_urls(), include_schemas_tab=True)
        spreadsheet_builder.save_workbook()

        os.remove(temp_yaml_filename)
        now = datetime.datetime.now()
        export_filename = "hca_spreadsheet-" + now.strftime("%Y-%m-%dT%H-%M-%S") + ".xlsx"

        response = make_response(ssheet_file.read())
        response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers.set('Content-Disposition', 'attachment',
                             filename=export_filename)
        os.remove(temp_filename)
        return response

# convenience function that processes an uploaded YAML file to identify which properties should be preselected
def _process_uploaded_file(file):
    selected_schemas = []
    # selected_references = []
    selected_properties = {}

    for schema in file:
        schema_name = list(schema.keys())[0]
        selected_schemas.append(schema_name)
        properties = schema[schema_name]['columns']

        props = []

        for prop in properties:
            props.append(prop)
        selected_properties[schema_name] = props

    return selected_schemas, selected_properties

# Convenience function that given a list of preselected schemas and properties,
# marks the appropriate schemas and properties as to be selected in the full schemas list.
# WARNING - does not include migrations, so if an old property is present, this will be added as a separate property!
def _preselect_properties(schema_properties, selected_schemas, selected_references, selected_properties):
    for schema in schema_properties:
        if schema["name"] in selected_schemas:
            schema["pre-selected"] = True
            properties = schema["properties"]

            if selected_references:
                for ref in selected_references:
                    t = ref.split(':')[0]
                    val = ref.split(':')[1]
                    if schema["name"] == t or schema["name"] == val:
                        for prop in properties.keys():
                            if (val == prop.split('.')[1] or (
                                    len(prop.split('.')) == 2 and prop.split('.')[0] != 'process')) and properties[
                                prop] == "not required":
                                schema["properties"][prop] = "pre-selected"
            if selected_properties:
                if schema["name"] in selected_properties:
                    sel_props = selected_properties[schema["name"]]

                    for prop in sel_props:
                        # Required properties will be pre-selected anyway so don't need to be marked again
                        if prop in list(properties.keys()) and properties[prop] == "not required":
                                schema["properties"][prop] = "pre-selected"
                        # This clause was intended to deal with linking properties etc but also lets through deleted and updated properties in this schema
                        # TO DO: implement migration lookup for this clause for properties that are in this schema
                        elif prop not in list(properties.keys()):
                            schema["properties"][prop] = "pre-selected"
    return schema_properties

# helper function to load the config file
def _loadConfig(file):
    config_file = configparser.ConfigParser(allow_no_value=True)
    config_file.read(file)
    return config_file

# flask convenience function
def _allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# helper function to convert properties as presented in the schema template library to a format readable by the generator UI
def _process_schemas():

    tab_config = SCHEMA_TEMPLATE.get_tabs_config()

    unordered = {}
    all_properties = []
    # go through the schemas from the tab config one by one
    for schema in tab_config.lookup('tabs'):
        property = {}

        schema_name = list(schema.keys())[0]

        # set the schema title, name and selection status
        property["title"] = schema[schema_name]["display_name"]
        property["name"] = schema_name
        property["select"] = False
        if "properties" not in property:
            property["properties"] = {}

        # go through all the properties in the schema
        for p in schema[schema_name]['columns']:
            # this deals with core and module properties
            if len(p.split(".")) > 2:
                parent = ".".join(p.split(".")[:-1])

                # special case for modules with ontology imports where the field is required if the module is used, eg donor.timecourse.unit.text
                if len(p.split(".")) == 4:
                    if SCHEMA_TEMPLATE.lookup(parent+".required"):
                        parent = ".".join(parent.split(".")[:-1])
                if SCHEMA_TEMPLATE.lookup(parent+".required"):
                    if SCHEMA_TEMPLATE.lookup(p + ".required"):
                        property["properties"][p] = "required"
                    else:
                        property["properties"][p] = "not required"
                else:
                    property["properties"][p]="not required"
            else:
                if SCHEMA_TEMPLATE.lookup(p+".required"):
                    property["properties"][p]="required"
                else:
                    property["properties"][p]="not required"

        # create a separate process object for appending to other properties below
        if property["name"] == "process":
            process = property["properties"]
            for k in process.keys():
                if process[k] == "required":
                    process[k] = "not required"

        unordered[property["name"]] = property

        # add the title of the schema to the global display name map
        DISPLAY_NAME_MAP[property["name"]] = property["title"]

    # add default biomaterial linking columns
    if 'biomaterial_linking' in CONFIG_FILE:
        for key in CONFIG_FILE['biomaterial_linking'].keys():
            if key in unordered.keys() and CONFIG_FILE['biomaterial_linking'][key] in unordered.keys():
                linking_field = list(unordered[CONFIG_FILE['biomaterial_linking'][key]]['properties'].keys())[0]
                unordered[key]['properties'][linking_field] = "not required"

    # add default protocol linking columns
    if 'protocol_linking' in CONFIG_FILE:
        for key in CONFIG_FILE['protocol_linking'].keys():
            if key in unordered.keys():
                protocols = []
                if "," in CONFIG_FILE['protocol_linking'][key]:
                    protocols = CONFIG_FILE['protocol_linking'][key].split(",")
                else:
                    protocols.append(CONFIG_FILE['protocol_linking'][key])

                for prot in protocols:
                    if prot in unordered.keys():
                        linking_field = list(unordered[prot]['properties'].keys())[0]
                        unordered[key]['properties'][linking_field] = "not required"

    # schemas should be ordered as per the ordering in the config file
    if 'ordering' in CONFIG_FILE:
        for key in CONFIG_FILE['ordering'].keys():
            if key in unordered.keys():
                # if the schema should have process fields appended according to the config file, append these
                if CONFIG_FILE['ordering'][key] == 'process' and process != '':
                    unordered[key]["properties"].update(process)

                all_properties.append(unordered[key])
            # deal with sub-tabs
            elif CONFIG_FILE['ordering'][key] != '':
                parent = CONFIG_FILE['ordering'][key]
                if parent in unordered.keys():
                    new_property = {}

                    new_property["title"] = tab_config.lookup('meta_data_properties')[parent][key]['user_friendly']

                    # tabs can't have a name that's longer than 32 characteres
                    if len(DISPLAY_NAME_MAP[parent] + " - " + new_property["title"]) < 32:
                        new_property["title"] = DISPLAY_NAME_MAP[parent] + " - " + new_property["title"]
                    new_property["name"] = key
                    new_property["select"] = False
                    if "properties" not in new_property:
                        new_property["properties"] = {}

                    for prop in unordered[parent]['properties']:
                        if key in prop:
                            new_property["properties"][prop] = unordered[parent]['properties'][prop]

                    for moved_prop in new_property["properties"]:
                        unordered[parent]['properties'].pop(moved_prop)

                    DISPLAY_NAME_MAP[new_property["name"]] = new_property["title"]

                    all_properties.append(new_property)

                    print(key + " is a recorded sub-property")
            else:
                print(key + " is currently not a recorded property")

    return all_properties

# helper function to extract the $ref properties (core and module references) from a schema
def _extract_references(properties, name, title, schema):

    # direct property = properties of the schema (incl wrapper properties for imports)
    direct_properties = []
    for property in properties:
        prop = property.split('.')[1]
        direct_properties.append(prop)

    structure = {}
    structure["title"] = title
    structure["name"] = name

    references = {}

    for dp in direct_properties:
        if dp not in EXCLUDED_PROPERTIES:
            # if the property exists and has a 'value_type' field and 'value_type' is an object and the property isn't already on the list of references
            if schema[dp] and schema[dp]['value_type'] and schema[dp]['value_type'] == 'object' and dp not in references.keys():
                if schema[dp]['required']:
                    references[dp] = "required"
                else:
                    references[dp] = "not required"
    structure["references"] = references
    return structure

# helper function to actually migrate the schema
def _migrate_schema(workbook, schema_url):
    schema_key = schema_url.split('/')[-1]
    schema_version = schema_url.split('/')[-2]

    linked_tabs = []

    # find out from the config file if this schema has dependent sub-schemas (eg contact, publications etc for project)
    if 'ordering' in CONFIG_FILE:
        if 'ordering' in CONFIG_FILE:
            for key in CONFIG_FILE['ordering'].keys():
                if CONFIG_FILE['ordering'][key] == schema_key:
                    linked_tabs.append(key)


    tab_config = SCHEMA_TEMPLATE.get_tabs_config()

    for schema in tab_config.lookup('tabs'):
        if schema_key == list(schema.keys())[0]:
            tab_name = schema[schema_key]['display_name']

            _update_tab(workbook, schema_key, tab_name, schema_version)

            # if there are dependent tabs, update these as well
            if linked_tabs:
                for tab in linked_tabs:
                    linked_tab_name = tab_config.lookup('meta_data_properties')[schema_key][tab]['user_friendly']

                    try:
                        workbook[linked_tab_name]
                    except:
                        print("No tab found for key " + linked_tab_name)

                        if len(tab_name + " - " + linked_tab_name) < 32:
                            linked_tab_name = tab_name + " - " + linked_tab_name
                        else:
                            linked_tab_name = linked_tab_name

                    _update_tab(workbook, schema_key, linked_tab_name, schema_version)

# convenience function to update a given tab in the work book
def _update_tab(workbook, schema_name, tab_name, schema_version):

    try:
        current_tab = workbook[tab_name]

        # look only a programmatic names in row 4
        for col in current_tab.iter_cols(min_row=4, max_row=4):
            for cell in col:

                # if there is actually a programmatic name in the cell, process this column
                if cell.value is not None:
                    # try to look up the programmatic name in the schema template - if success, update the user friendly properties as we don't know if the
                    # minor or patch schema version might have changed
                    try:
                        new_property = SCHEMA_TEMPLATE.lookup(cell.value)
                        _update_user_properties(cell.value, cell.col_idx, current_tab, SCHEMA_TEMPLATE.lookup(cell.value + ".required"), schema_name)

                    # if the property from the spreadsheet isn't found in the lookup, try to migrate it
                    except UnknownKeyException:
                        try:
                            new_property = SCHEMA_TEMPLATE.replaced_by_latest(cell.value)
                            # if a new property exists for the cell value, set the cell value to the new property, then update the user friendly fields for the column
                            if new_property is not "":
                                cell.value = new_property
                                _update_user_properties(new_property, cell.col_idx, current_tab, SCHEMA_TEMPLATE.lookup(new_property + ".required"), schema_name)

                        # if the migration lookup fails but there is a version at which this property was migrated, assume it was deleted and delete the column
                        except UnknownKeyException:
                            if SCHEMA_TEMPLATE._lookup_migration_version(cell.value) is not None:
                                current_tab.delete_cols(cell.col_idx, 1)
    except Exception:
       print("No tab found for key " + tab_name)




# WARNING: This code duplicates a large section of the ingest-client spreadsheet builder. Updates to the relevant spreadsheet builder code should be mimicked here!
def _update_user_properties(col_name, col_index, current_tab, required, tab_schema):

    # slightly complicated set-up to deal with the fact that any .text field from an ontology module should actually use the
    # user-friendly name/description/example/required status of its wrapper property
    if col_name.split(".")[-1] == "text":
        uf = _get_value_for_column(col_name.replace('.text', ''), "user_friendly").upper()
    else:
        uf = _get_value_for_column(col_name, "user_friendly").upper()
    if col_name.split(".")[-1] == "text":
        desc = _get_value_for_column(col_name.replace('.text', ''), "description")
        if desc == "":
            desc = _get_value_for_column(col_name, "description")
    else:
        desc = _get_value_for_column(col_name, "description")
    if col_name.split(".")[-1] == "text":
        required = bool(_get_value_for_column(col_name.replace('.text', ''), "required"))
    else:
        required = bool(_get_value_for_column(col_name, "required"))
    if col_name.split(".")[-1] == "text":
        example_text = _get_value_for_column(col_name.replace('.text', ''), "example")
        if example_text == "":
            example_text = _get_value_for_column(col_name, "example")
    else:
        example_text = _get_value_for_column(col_name, "example")
    if col_name.split(".")[-1] == "text":
        guidelines = _get_value_for_column(col_name.replace('.text', ''), "guidelines")
        if guidelines == "":
            guidelines = _get_value_for_column(col_name, "guidelines")
    else:
        guidelines = _get_value_for_column(col_name, "guidelines")

    # alternative set-up to deal with correct labelling of barcode and kit imports (eg UMI barcode - foo vs cell barcode - foo)
    wrapper = ".".join(col_name.split(".")[:-1])
    if SCHEMA_TEMPLATE.lookup(wrapper)['schema']['module'] \
            and (SCHEMA_TEMPLATE.lookup(wrapper)['schema']['module'] == 'purchased_reagents'
                 or SCHEMA_TEMPLATE.lookup(wrapper)['schema']['module'] == 'barcode') \
            and SCHEMA_TEMPLATE.lookup(wrapper)['multivalue'] == False:
        uf = (SCHEMA_TEMPLATE.lookup(wrapper)['user_friendly'] + " - " + uf).upper()

    # core fields like ID and name for biomaterial and protocol should be referred to by the schema name, eg donor organism id rather than biomaterial id
    if "BIOMATERIAL " in uf:
        schema_name = col_name.split(".")[0]

        for schema in SCHEMA_TEMPLATE.get_tabs_config().lookup('tabs'):
            if schema_name == list(schema.keys())[0]:
                schema_uf = schema[schema_name]['display_name']
        uf = uf.replace("BIOMATERIAL", schema_uf.upper())

        if tab_schema != schema_name:
            uf = "INPUT " + uf

    if "PROTOCOL " in uf:
        schema_name = col_name.split(".")[0]

        for schema in SCHEMA_TEMPLATE.get_tabs_config().lookup('tabs'):
            if schema_name == list(schema.keys())[0]:
                schema_uf = schema[schema_name]['display_name']
        uf = uf.replace("PROTOCOL", schema_uf.upper())

    if required:
        uf = uf + " (Required)"

    current_tab.cell(row=1, column=col_index, value=uf)
    # set the description
    current_tab.cell(row=2, column=col_index, value=desc)

    # write example
    if example_text:
        current_tab.cell(row=3, column=col_index, value=guidelines + ' For example: ' + example_text)
    else:
        current_tab.cell(row=3, column=col_index, value=guidelines)

# convenience function to look up a property (user friendly name, description etc) for a column from the schema template library
def _get_value_for_column(col_name, property):
    try:
        uf = str(SCHEMA_TEMPLATE.lookup(col_name + "." + property)) if SCHEMA_TEMPLATE.lookup(col_name + "." + property) else ""
        return uf
    except Exception:
        print("No property " + property + " for " + col_name)
        return ""


# main launcher for running the app locally
if __name__ == '__main__':

    logging.basicConfig(stream=sys.stdout, level=logging.INFO)

    # if '/generator' in dir:
    #     dir = dir.replace('/generator', '')
    # base_uri = dir + "/"


    CONFIG_FILE = _loadConfig('config.ini')

    env = ''
    if 'system' in CONFIG_FILE and 'environment' in CONFIG_FILE['system']:
        env = CONFIG_FILE['system']['environment']

    if env == 'prod':
        api_url = INGEST_API_URL.replace("{env}.", '')
    else:
        api_url = INGEST_API_URL.replace("{env}", env)

    SCHEMA_TEMPLATE = SchemaTemplate(ingest_api_url=api_url,migrations_url='https://schema.dev.data.humancellatlas.org/property_migrations')

    app.run(host='0.0.0.0', port=5000)